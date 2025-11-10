# gestion/views/movement_views.py

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.utils.html import format_html_join
from django.views.generic import ListView
from django.contrib.auth.mixins import UserPassesTestMixin
from django.contrib.auth.decorators import user_passes_test, login_required
from django.db.models import Q
from django.template.loader import render_to_string
import json
import io
import openpyxl

# --- Importaciones de Seguridad ---
from ..auth_utils import is_bodega_or_admin # Importamos nuestro chequeo de rol

from ..forms.movement_forms import ProductMovementForm
from ..models import Inventory, ProductMovement

# --- APLICAMOS SEGURIDAD DE ROL 'BODEGA' O 'ADMIN' ---
@user_passes_test(is_bodega_or_admin, login_url='login')
@require_http_methods(["GET", "POST"])
def product_movement_view(request):
    
    if request.method == 'GET':
        form = ProductMovementForm()
        return render(request, 'gestion/movement_form.html', {'form': form})

    # Si es POST (petición AJAX)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'errors': 'JSON inválido'}, status=400)

    form = ProductMovementForm(data)

    if not form.is_valid():
        # Formateamos los errores para SweetAlert2
        errors_html = format_html_join(
            '<br>', 
            '{}',
            ((error,) for field, error_list in form.errors.items() for error in error_list)
        )
        return JsonResponse({'status': 'error', 'errors': errors_html}, status=400)

    # Si el formulario es VÁLIDO
    try:
        with transaction.atomic():
            movement = form.save(commit=False)
            movement.performed_by = request.user 
            movement.movement_type = 'TRANSFER' # Hard-codeamos el tipo
            movement.save()

            quantity = movement.quantity
            
            # 1. Resta del origen
            origin_stock = Inventory.objects.get(
                product=movement.product, 
                zone=movement.origin_zone
            )
            origin_stock.quantity -= quantity
            origin_stock.save()

            # 2. Suma al destino
            destination_stock, created = Inventory.objects.get_or_create(
                product=movement.product,
                zone=movement.destination_zone,
                defaults={'quantity': 0}
            )
            destination_stock.quantity += quantity
            destination_stock.save()

        return JsonResponse({
            'status': 'success', 
            'message': 'Transferencia realizada con éxito.'
        })

    except Exception as e:
        # Captura cualquier error inesperado
        return JsonResponse({'status': 'error', 'errors': f'Error del servidor: {str(e)}'}, status=500)


# ----------------------------------------------------
# VISTA PARA LISTAR MOVIMIENTOS
# ----------------------------------------------------
class ProductMovementListView(UserPassesTestMixin, ListView):
    model = ProductMovement
    template_name = 'gestion/movement_list.html'
    context_object_name = 'movements'
    paginate_by = 20
    
    def test_func(self):
        return is_bodega_or_admin(self.request.user)
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('product', 'origin_zone', 'destination_zone', 'performed_by')
        
        search_query = self.request.GET.get('q', None)
        if search_query:
            queryset = queryset.filter(
                Q(product__name__icontains=search_query) |
                Q(product__sku__icontains=search_query) |
                Q(reason__icontains=search_query)
            )
        
        movement_type = self.request.GET.get('type', None)
        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)
        
        # Guardar filtros en sesión
        if 'q' in self.request.GET:
            self.request.session['movement_search_query'] = search_query
        elif 'movement_search_query' in self.request.session:
            pass  # Ya se aplicó arriba
        
        if 'type' in self.request.GET:
            self.request.session['movement_type_filter'] = movement_type
        elif 'movement_type_filter' in self.request.session:
            pass
        
        # Ordenamiento
        sort_by = self.request.GET.get('sort', '-movement_date')
        sort_options = {
            'date': '-movement_date',  # Fecha más reciente primero
            '-date': 'movement_date',  # Fecha más antigua primero
            'product': 'product__name',  # Producto A-Z
            '-product': '-product__name',  # Producto Z-A
            'quantity': 'quantity',  # Cantidad menor a mayor
            '-quantity': '-quantity',  # Cantidad mayor a menor
            'type': 'movement_type',  # Tipo A-Z
            '-type': '-movement_type',  # Tipo Z-A
        }
        
        if sort_by in sort_options:
            queryset = queryset.order_by(sort_options[sort_by])
        else:
            queryset = queryset.order_by('-movement_date')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('q', '')
        if 'movement_search_query' in self.request.session and not context['search_query']:
            context['search_query'] = self.request.session['movement_search_query']
        
        context['type_filter'] = self.request.GET.get('type', '')
        if 'movement_type_filter' in self.request.session and not context['type_filter']:
            context['type_filter'] = self.request.session['movement_type_filter']
        
        context['current_sort'] = self.request.GET.get('sort', '-movement_date')
        
        return context


# ----------------------------------------------------
# EXPORTAR MOVIMIENTOS A EXCEL
# ----------------------------------------------------
@login_required
def export_movements_excel(request):
    """
    Crea un archivo Excel con la lista de movimientos.
    Respeta los filtros de búsqueda y tipo si están presentes.
    """
    if not is_bodega_or_admin(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("No tienes permisos para exportar movimientos")
    
    # Obtener filtros de la URL o de la sesión
    search_query = request.GET.get('q', '')
    if not search_query and 'movement_search_query' in request.session:
        search_query = request.session.get('movement_search_query', '')
    
    movement_type = request.GET.get('type', '')
    if not movement_type and 'movement_type_filter' in request.session:
        movement_type = request.session.get('movement_type_filter', '')
    
    # Construir queryset con los mismos filtros que la vista de lista
    movements = ProductMovement.objects.all().select_related(
        'product', 'origin_zone', 'destination_zone', 'performed_by',
        'origin_zone__warehouse', 'destination_zone__warehouse'
    )
    
    # Aplicar filtros
    if search_query:
        movements = movements.filter(
            Q(product__name__icontains=search_query) |
            Q(product__sku__icontains=search_query) |
            Q(reason__icontains=search_query)
        )
    
    if movement_type:
        movements = movements.filter(movement_type=movement_type)
    
    movements = movements.order_by('-movement_date')

    # Crear el workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Movimientos"

    # Encabezados
    headers = ['Producto', 'SKU', 'Cantidad', 'Tipo', 'Bodega Origen', 'Zona Origen', 'Bodega Destino', 'Zona Destino', 'Fecha', 'Realizado por', 'Motivo']
    ws.append(headers)

    # Estilo para encabezados
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar datos
    for movement in movements:
        origin_warehouse = movement.origin_zone.warehouse.name if movement.origin_zone and movement.origin_zone.warehouse else 'N/A'
        destination_warehouse = movement.destination_zone.warehouse.name if movement.destination_zone and movement.destination_zone.warehouse else 'N/A'
        origin_zone = movement.origin_zone.name if movement.origin_zone else 'N/A'
        destination_zone = movement.destination_zone.name if movement.destination_zone else 'N/A'
        
        ws.append([
            movement.product.name,
            movement.product.sku,
            movement.quantity,
            movement.get_movement_type_display(),
            origin_warehouse,
            origin_zone,
            destination_warehouse,
            destination_zone,
            movement.movement_date.strftime('%d/%m/%Y %H:%M') if movement.movement_date else 'N/A',
            movement.performed_by.get_full_name() or movement.performed_by.username if movement.performed_by else 'N/A',
            movement.reason or 'N/A'
        ])

    # Ajustar ancho de columnas
    column_widths = {
        'A': 30,  # Producto
        'B': 20,  # SKU
        'C': 12,  # Cantidad
        'D': 20,  # Tipo
        'E': 20,  # Bodega Origen
        'F': 20,  # Zona Origen
        'G': 20,  # Bodega Destino
        'H': 20,  # Zona Destino
        'I': 18,  # Fecha
        'J': 20,  # Realizado por
        'K': 30,  # Motivo
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Generar nombre del archivo con filtros si existen
    filename = "lista_movimientos.xlsx"
    if search_query or movement_type:
        parts = ["lista_movimientos"]
        if movement_type:
            parts.append(movement_type.lower())
        if search_query:
            parts.append("filtrado")
        filename = "_".join(parts) + ".xlsx"

    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)

    response = HttpResponse(
        virtual_workbook.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# ----------------------------------------------------
# VISTA AJAX PARA LIVE SEARCH DE MOVIMIENTOS
# ----------------------------------------------------
@login_required
def search_movements_live(request):
    """
    Busca movimientos por 'q' y 'type', y devuelve el HTML de la tabla.
    """
    if not is_bodega_or_admin(request.user):
        return JsonResponse({'html': '<tr><td colspan="9" class="text-center text-danger">No tienes permisos.</td></tr>'})
    
    search_query = request.GET.get('q', '')
    movement_type = request.GET.get('type', '')
    
    queryset = ProductMovement.objects.all().select_related('product', 'origin_zone', 'destination_zone', 'performed_by')
    
    if search_query:
        queryset = queryset.filter(
            Q(product__name__icontains=search_query) |
            Q(product__sku__icontains=search_query) |
            Q(reason__icontains=search_query)
        )
    
    if movement_type:
        queryset = queryset.filter(movement_type=movement_type)
    
    # Ordenamiento
    sort_by = request.GET.get('sort', '-movement_date')
    sort_options = {
        'date': '-movement_date',
        '-date': 'movement_date',
        'product': 'product__name',
        '-product': '-product__name',
        'quantity': 'quantity',
        '-quantity': '-quantity',
        'type': 'movement_type',
        '-type': '-movement_type',
    }
    
    if sort_by in sort_options:
        queryset = queryset.order_by(sort_options[sort_by])
    else:
        queryset = queryset.order_by('-movement_date')
    
    context = {
        'movements': queryset,
    }

    html_results = render_to_string(
        'gestion/includes/movement_table_body.html', 
        context, 
        request=request
    )
    return JsonResponse({'html': html_results})