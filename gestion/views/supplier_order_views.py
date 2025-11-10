# gestion/views/supplier_order_views.py

from django.views.generic import ListView, CreateView, DetailView
from django.contrib.auth.mixins import UserPassesTestMixin
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.utils import timezone
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
import io
import openpyxl

from ..models import SupplierOrder, SupplierOrderItem, Inventory, ProductMovement
from ..forms.supplier_order_forms import SupplierOrderForm, SupplierOrderItemForm
from ..auth_utils import is_admin, is_bodega_or_admin


class SupplierOrderListView(UserPassesTestMixin, ListView):
    model = SupplierOrder
    template_name = 'gestion/supplier_order_list.html'
    context_object_name = 'orders'
    paginate_by = 20
    
    def test_func(self):
        return is_admin(self.request.user) or is_bodega_or_admin(self.request.user)
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('supplier', 'warehouse', 'zone', 'requested_by')
        
        status_filter = self.request.GET.get('status', None)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        search_query = self.request.GET.get('q', None)
        if search_query:
            queryset = queryset.filter(
                Q(supplier__name__icontains=search_query) |
                Q(id=search_query) if search_query.isdigit() else Q()
            )
        
        # Ordenamiento
        sort_by = self.request.GET.get('sort', '-order_date')
        sort_options = {
            'date': '-order_date',  # Fecha más reciente primero
            '-date': 'order_date',  # Fecha más antigua primero
            'supplier': 'supplier__name',  # Proveedor A-Z
            '-supplier': '-supplier__name',  # Proveedor Z-A
            'status': 'status',  # Estado A-Z
            '-status': '-status',  # Estado Z-A
            'id': 'id',  # ID menor a mayor
            '-id': '-id',  # ID mayor a menor
        }
        
        if sort_by in sort_options:
            queryset = queryset.order_by(sort_options[sort_by])
        else:
            queryset = queryset.order_by('-order_date')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('q', '')
        context['status_filter'] = self.request.GET.get('status', '')
        context['current_sort'] = self.request.GET.get('sort', '-order_date')
        
        # Guardar filtros en sesión
        if 'q' in self.request.GET:
            self.request.session['supplier_order_search_query'] = self.request.GET.get('q', '')
        elif 'supplier_order_search_query' in self.request.session:
            context['search_query'] = self.request.session['supplier_order_search_query']
        
        if 'status' in self.request.GET:
            self.request.session['supplier_order_status_filter'] = self.request.GET.get('status', '')
        elif 'supplier_order_status_filter' in self.request.session:
            context['status_filter'] = self.request.session['supplier_order_status_filter']
        
        return context


@login_required
def supplier_order_create(request):
    """
    Vista para crear una nueva orden a proveedor.
    """
    if not (is_admin(request.user) or is_bodega_or_admin(request.user)):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("No tienes permisos para crear órdenes")
    
    if request.method == 'POST':
        form = SupplierOrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            order.requested_by = request.user
            order.save()
            return redirect('supplier_order_add_items', pk=order.pk)
    else:
        form = SupplierOrderForm()
    
    return render(request, 'gestion/supplier_order_form.html', {
        'form': form,
        'title': 'Nueva Orden a Proveedor'
    })


@login_required
def supplier_order_add_items(request, pk):
    """
    Vista para agregar items a una orden con catálogo visual.
    """
    if not (is_admin(request.user) or is_bodega_or_admin(request.user)):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("No tienes permisos para editar órdenes")
    
    order = get_object_or_404(SupplierOrder, pk=pk)
    
    if order.status != 'PENDING':
        messages.error(request, 'No se pueden modificar órdenes que no están pendientes.')
        return redirect('supplier_order_detail', pk=order.pk)
    
    # Obtener productos del proveedor de la orden
    from ..models import Product
    products = Product.objects.filter(
        supplier=order.supplier,
        is_active=True
    ).select_related('supplier').order_by('name')
    
    # Si hay búsqueda, filtrar productos
    search_query = request.GET.get('q', '')
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(sku__icontains=search_query)
        )
    
    return render(request, 'gestion/supplier_order_add_items.html', {
        'order': order,
        'products': products,
        'items': order.items.select_related('product').all(),
        'search_query': search_query
    })


class SupplierOrderDetailView(UserPassesTestMixin, DetailView):
    model = SupplierOrder
    template_name = 'gestion/supplier_order_detail.html'
    context_object_name = 'order'
    
    def test_func(self):
        return is_admin(self.request.user) or is_bodega_or_admin(self.request.user)
    
    def get_queryset(self):
        return super().get_queryset().select_related(
            'supplier', 'warehouse', 'zone', 'requested_by'
        ).prefetch_related('items__product')


@login_required
def supplier_order_receive(request, pk):
    """
    Vista para marcar una orden como recibida y actualizar el inventario.
    """
    if not (is_admin(request.user) or is_bodega_or_admin(request.user)):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("No tienes permisos para recibir órdenes")
    
    order = get_object_or_404(SupplierOrder, pk=pk)
    
    if order.status != 'PENDING':
        messages.error(request, 'Solo se pueden recibir órdenes pendientes.')
        return redirect('supplier_order_detail', pk=order.pk)
    
    if order.items.count() == 0:
        messages.error(request, 'La orden no tiene items. Agrega productos antes de recibirla.')
        return redirect('supplier_order_add_items', pk=order.pk)
    
    if request.method == 'POST':
        with transaction.atomic():
            # Actualizar el estado de la orden
            order.status = 'RECEIVED'
            order.received_date = timezone.now()
            order.save()
            
            # Procesar cada item
            for item in order.items.select_related('product').all():
                # Crear o actualizar inventario
                inventory, created = Inventory.objects.get_or_create(
                    product=item.product,
                    zone=order.zone,
                    defaults={'quantity': 0}
                )
                inventory.quantity += item.quantity
                inventory.save()
                
                # Crear movimiento de entrada
                ProductMovement.objects.create(
                    product=item.product,
                    quantity=item.quantity,
                    movement_type='ENTRY',
                    destination_zone=order.zone,
                    performed_by=request.user,
                    reason=f"Recibo de orden #{order.id} de {order.supplier.name}"
                )
            
            # Verificar si algún producto pasó de stock 0 a stock > 0
            products_restocked = []
            for item in order.items.select_related('product').all():
                if item.product.total_quantity > 0:
                    products_restocked.append(item.product.name)
            
            if products_restocked:
                messages.success(
                    request, 
                    f'Orden #{order.id} recibida exitosamente. Inventario actualizado. '
                    f'Los productos ahora aparecerán en la lista de inventario.'
                )
            else:
                messages.success(request, f'Orden #{order.id} recibida exitosamente. Inventario actualizado.')
            
            return redirect('supplier_order_detail', pk=order.pk)
    
    return render(request, 'gestion/supplier_order_receive.html', {
        'order': order
    })


@login_required
def supplier_order_item_delete(request, order_pk, item_pk):
    """
    Vista para eliminar un item de una orden.
    """
    if not (is_admin(request.user) or is_bodega_or_admin(request.user)):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("No tienes permisos para eliminar items")
    
    order = get_object_or_404(SupplierOrder, pk=order_pk)
    item = get_object_or_404(SupplierOrderItem, pk=item_pk, order=order)
    
    if order.status != 'PENDING':
        messages.error(request, 'No se pueden modificar órdenes que no están pendientes.')
        return redirect('supplier_order_detail', pk=order.pk)
    
    if request.method == 'POST':
        item_name = item.product.name
        item.delete()
        messages.success(request, f'Item "{item_name}" eliminado de la orden exitosamente.')
        return redirect('supplier_order_add_items', pk=order.pk)
    
    return render(request, 'gestion/supplier_order_item_delete.html', {
        'order': order,
        'item': item
    })


# ----------------------------------------------------
# VISTA AJAX PARA LIVE SEARCH DE ÓRDENES A PROVEEDOR
# ----------------------------------------------------
@login_required
def search_supplier_orders_live_list(request):
    """
    Busca órdenes a proveedor por 'q' y 'status', y devuelve el HTML de la tabla.
    """
    if not (is_admin(request.user) or is_bodega_or_admin(request.user)):
        return JsonResponse({'html': '<tr><td colspan="8" class="text-center text-danger">No tienes permisos.</td></tr>'})
    
    search_query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    
    queryset = SupplierOrder.objects.select_related('supplier', 'warehouse', 'zone', 'requested_by')
    
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    
    if search_query:
        if search_query.isdigit():
            queryset = queryset.filter(id=search_query)
        else:
            queryset = queryset.filter(supplier__name__icontains=search_query)
    
    queryset = queryset.order_by('-order_date')
    
    context = {
        'orders': queryset,
    }

    html_results = render_to_string(
        'gestion/includes/supplier_order_table_body.html', 
        context, 
        request=request
    )
    return JsonResponse({'html': html_results})


# ----------------------------------------------------
# EXPORTAR ÓRDENES A PROVEEDOR A EXCEL
# ----------------------------------------------------
@login_required
def export_supplier_orders_excel(request):
    """
    Crea un archivo Excel con el historial de órdenes a proveedor.
    Respeta los filtros de búsqueda y estado si están presentes.
    """
    if not (is_admin(request.user) or is_bodega_or_admin(request.user)):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("No tienes permisos para exportar órdenes")
    
    # Obtener filtros de la URL o de la sesión
    search_query = request.GET.get('q', '')
    if not search_query and 'supplier_order_search_query' in request.session:
        search_query = request.session.get('supplier_order_search_query', '')
    
    status_filter = request.GET.get('status', '')
    if not status_filter and 'supplier_order_status_filter' in request.session:
        status_filter = request.session.get('supplier_order_status_filter', '')
    
    # Construir queryset con los mismos filtros que la vista de lista
    orders = SupplierOrder.objects.all().select_related(
        'supplier', 'warehouse', 'zone', 'requested_by'
    ).prefetch_related('items__product')
    
    # Aplicar filtros
    if status_filter:
        orders = orders.filter(status=status_filter)
    
    if search_query:
        if search_query.isdigit():
            orders = orders.filter(id=search_query)
        else:
            orders = orders.filter(supplier__name__icontains=search_query)
    
    orders = orders.order_by('-order_date')

    # Crear el workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial de Órdenes a Proveedor"

    # Encabezados
    headers = ['ID Orden', 'Proveedor', 'Bodega', 'Zona', 'Items', 'Cantidad Total', 'Fecha', 'Fecha Recepción', 'Estado', 'Solicitado por']
    ws.append(headers)

    # Estilo para encabezados
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar datos
    for order in orders:
        ws.append([
            order.id,
            order.supplier.name if order.supplier else 'N/A',
            order.warehouse.name if order.warehouse else 'N/A',
            order.zone.name if order.zone else 'N/A',
            order.total_items,
            order.total_quantity,
            order.order_date.strftime('%d/%m/%Y %H:%M') if order.order_date else 'N/A',
            order.received_date.strftime('%d/%m/%Y %H:%M') if order.received_date else 'N/A',
            order.get_status_display(),
            order.requested_by.get_full_name() or order.requested_by.username if order.requested_by else 'N/A'
        ])

    # Ajustar ancho de columnas
    column_widths = {
        'A': 12,  # ID Orden
        'B': 25,  # Proveedor
        'C': 20,  # Bodega
        'D': 20,  # Zona
        'E': 10,  # Items
        'F': 15,  # Cantidad Total
        'G': 18,  # Fecha
        'H': 18,  # Fecha Recepción
        'I': 15,  # Estado
        'J': 20,  # Solicitado por
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Generar nombre del archivo con filtros si existen
    filename = "historial_ordenes_proveedor.xlsx"
    if search_query or status_filter:
        parts = ["historial_ordenes_proveedor"]
        if status_filter:
            parts.append(status_filter.lower())
        if search_query:
            parts.append("filtrado")
        filename = "_".join(parts) + ".xlsx"

    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)

    response = HttpResponse(
        virtual_workbook.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

