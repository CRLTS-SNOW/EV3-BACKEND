# gestion/views/user_views.py

import io
import openpyxl
from django.views.generic import ListView, CreateView, UpdateView
from django.contrib.auth.mixins import UserPassesTestMixin
from django.contrib.auth.models import User
from django.contrib.auth.decorators import user_passes_test, login_required
from django.urls import reverse_lazy
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string

from ..models import UserProfile
from ..forms.user_forms import UserCreateForm, UserUpdateForm, UserPasswordChangeForm
from ..auth_utils import is_admin
from gestion.firebase_service import sync_django_user_to_firebase, get_firebase_user_by_email, delete_firebase_user


class UserListView(UserPassesTestMixin, ListView):
    model = User
    template_name = 'gestion/user_list.html'
    context_object_name = 'users'
    paginate_by = 20
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('profile')
        
        search_query = self.request.GET.get('q', None)
        if search_query:
            queryset = queryset.filter(
                Q(username__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query)
            )
        
        # Ordenamiento
        sort_by = self.request.GET.get('sort', 'username')
        sort_options = {
            'username': 'username',  # Usuario A-Z
            '-username': '-username',  # Usuario Z-A
            'email': 'email',  # Email A-Z
            '-email': '-email',  # Email Z-A
            'first_name': 'first_name',  # Nombre A-Z
            '-first_name': '-first_name',  # Nombre Z-A
            'last_name': 'last_name',  # Apellido A-Z
            '-last_name': '-last_name',  # Apellido Z-A
        }
        
        if sort_by in sort_options:
            queryset = queryset.order_by(sort_options[sort_by])
        else:
            queryset = queryset.order_by('username')
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('q', '')
        context['current_sort'] = self.request.GET.get('sort', 'username')
        
        # Guardar búsqueda en sesión
        if 'q' in self.request.GET:
            self.request.session['user_search_query'] = self.request.GET.get('q', '')
        elif 'user_search_query' in self.request.session:
            context['search_query'] = self.request.session['user_search_query']
        
        return context


# ----------------------------------------------------
# EXPORTAR USUARIOS A EXCEL
# ----------------------------------------------------
@login_required
def export_users_excel(request):
    """
    Crea un archivo Excel con la lista de usuarios.
    Respeta los filtros de búsqueda si están presentes.
    """
    if not is_admin(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("No tienes permisos para exportar usuarios")
    
    # Obtener filtros de la URL o de la sesión
    search_query = request.GET.get('q', '')
    if not search_query and 'user_search_query' in request.session:
        search_query = request.session.get('user_search_query', '')
    
    # Construir queryset con los mismos filtros que la vista de lista
    users = User.objects.all().select_related('profile', 'profile__warehouse')
    
    # Aplicar filtros
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
    
    users = users.order_by('username')

    # Crear el workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Usuarios"

    # Encabezados
    headers = ['Usuario', 'Nombre', 'Apellido', 'Nombre Completo', 'Email', 'Teléfono', 'Rol', 'Bodega Asignada', 'Estado', 'Fecha de Creación']
    ws.append(headers)

    # Estilo para encabezados
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar datos
    for user in users:
        profile = user.profile if hasattr(user, 'profile') else None
        full_name = user.get_full_name() or ''
        role = profile.get_role_display() if profile and profile.role else 'N/A'
        warehouse = profile.warehouse.name if profile and profile.warehouse else 'N/A'
        phone = profile.phone if profile and profile.phone else 'N/A'
        status = 'Activo' if user.is_active else 'Inactivo'
        date_joined = user.date_joined.strftime('%d/%m/%Y %H:%M') if user.date_joined else 'N/A'
        
        ws.append([
            user.username,
            user.first_name or '',
            user.last_name or '',
            full_name,
            user.email or 'N/A',
            phone,
            role,
            warehouse,
            status,
            date_joined
        ])

    # Ajustar ancho de columnas
    column_widths = {
        'A': 20,  # Usuario
        'B': 20,  # Nombre
        'C': 20,  # Apellido
        'D': 25,  # Nombre Completo
        'E': 30,  # Email
        'F': 15,  # Teléfono
        'G': 20,  # Rol
        'H': 25,  # Bodega Asignada
        'I': 12,  # Estado
        'J': 18,  # Fecha de Creación
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Generar nombre del archivo con filtros si existen
    filename = "lista_usuarios.xlsx"
    if search_query:
        filename = "lista_usuarios_filtrado.xlsx"

    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)

    response = HttpResponse(
        virtual_workbook.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


class UserCreateView(UserPassesTestMixin, CreateView):
    model = User
    form_class = UserCreateForm
    template_name = 'gestion/user_form.html'
    success_url = reverse_lazy('user_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear Usuario'
        context['is_create'] = True
        return context
    
    def form_valid(self, form):
        response = super().form_valid(form)
        user = form.instance
        
        # Sincronizar AUTOMÁTICAMENTE con Firebase si tiene email
        # La función sync_django_user_to_firebase crea el usuario en Firebase
        # incluso si no se proporciona contraseña (usa contraseña temporal)
        from gestion.firebase_service import sync_django_user_to_firebase
        
        if user.email:
            password = form.cleaned_data.get('password')
            # Sincronizar con Firebase (si hay contraseña, la usa; si no, crea con temporal)
            firebase_user = sync_django_user_to_firebase(user, password=password, old_email=None)
            
            if firebase_user:
                if password:
                    messages.success(self.request, f'Usuario {user.username} creado y sincronizado con Firebase exitosamente.')
                else:
                    messages.success(self.request, f'Usuario {user.username} creado en Django y Firebase. El usuario deberá usar "Olvidé mi contraseña" para establecer una contraseña.')
            else:
                messages.warning(self.request, f'Usuario {user.username} creado en Django, pero no se pudo sincronizar con Firebase. Verifica la configuración.')
        else:
            messages.warning(self.request, f'Usuario {user.username} creado, pero no se puede sincronizar con Firebase sin email.')
        
        return response


class UserUpdateView(UserPassesTestMixin, UpdateView):
    model = User
    form_class = UserUpdateForm
    template_name = 'gestion/user_form.html'
    success_url = reverse_lazy('user_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Editar Usuario'
        context['is_create'] = False
        if self.object:
            context['user'] = self.object
        return context
    
    def form_valid(self, form):
        # Obtener el email anterior ANTES de guardar
        user = self.get_object()
        old_email = user.email if user.email else None
        
        # Guardar en Django
        response = super().form_valid(form)
        user = form.instance
        new_email = user.email if user.email else None
        
        # Sincronizar AUTOMÁTICAMENTE con Firebase
        # La función sync_django_user_to_firebase maneja todos los casos:
        # - Si el usuario existe en Firebase (por email actual o anterior), lo actualiza
        # - Si el usuario no existe, lo crea con contraseña temporal
        # - Si el email cambió, lo actualiza en Firebase
        from gestion.firebase_service import sync_django_user_to_firebase
        
        if new_email:
            # Sincronizar con Firebase (pasa el email anterior para búsqueda inteligente)
            firebase_user = sync_django_user_to_firebase(user, password=None, old_email=old_email)
            
            if firebase_user:
                email_changed = old_email and old_email.strip().lower() != new_email.strip().lower()
                if email_changed:
                    messages.success(self.request, f'Usuario {user.username} actualizado. Email sincronizado con Firebase: {old_email} → {new_email}')
                else:
                    messages.success(self.request, f'Usuario {user.username} actualizado y sincronizado con Firebase exitosamente.')
            else:
                # Si no se pudo sincronizar, informar al usuario
                # Esto puede pasar si Firebase no está configurado o hay un error
                messages.warning(self.request, f'Usuario {user.username} actualizado en Django, pero no se pudo sincronizar con Firebase. Verifica la configuración.')
        else:
            # Usuario sin email - no se puede sincronizar con Firebase
            messages.warning(self.request, f'Usuario {user.username} actualizado, pero no se puede sincronizar con Firebase sin email.')
        
        return response


@login_required
def user_change_own_password(request):
    """
    Vista para que un usuario cambie su propia contraseña.
    """
    user = request.user
    
    if request.method == 'POST':
        form = UserPasswordChangeForm(request.POST, user=user, is_own_password=True)
        if form.is_valid():
            new_password = form.cleaned_data['new_password']
            
            # Actualizar contraseña SOLO en Firebase (no en Django)
            if user.email:
                from gestion.firebase_service import get_firebase_user_by_email, update_firebase_user
                firebase_user = get_firebase_user_by_email(user.email)
                if firebase_user:
                    # Actualizar solo en Firebase
                    update_result = update_firebase_user(firebase_user.uid, password=new_password)
                    if update_result:
                        # NO actualizar la contraseña en Django - solo usar Firebase
                        # user.set_password(new_password)  # COMENTADO - solo Firebase
                        # user.save()  # COMENTADO - solo Firebase
                        messages.success(request, 'Tu contraseña ha sido actualizada en Firebase exitosamente.')
                    else:
                        messages.error(request, 'Error al actualizar la contraseña en Firebase.')
                else:
                    messages.error(request, 'Usuario no encontrado en Firebase. No se pudo actualizar la contraseña.')
            else:
                messages.error(request, 'No tienes email configurado. No se puede actualizar la contraseña en Firebase.')
            
            return redirect('product_list')
    else:
        form = UserPasswordChangeForm(user=user, is_own_password=True)
    
    return render(request, 'gestion/user_password_form.html', {
        'form': form,
        'user': user,
        'is_own_password': True
    })


@user_passes_test(is_admin, login_url='login')
def user_change_password(request, pk):
    """
    Vista para cambiar la contraseña de un usuario.
    Solo el admin puede cambiar contraseñas de otros usuarios.
    """
    user = get_object_or_404(User, pk=pk)
    is_own_password = (request.user == user)
    
    # Solo permitir que usuarios cambien su propia contraseña o que admin cambie cualquier contraseña
    if not is_own_password and not is_admin(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("No tienes permisos para cambiar esta contraseña.")
    
    if request.method == 'POST':
        form = UserPasswordChangeForm(request.POST, user=user, is_own_password=is_own_password)
        if form.is_valid():
            new_password = form.cleaned_data['new_password']
            
            # Actualizar contraseña SOLO en Firebase (no en Django)
            if user.email:
                from gestion.firebase_service import get_firebase_user_by_email, update_firebase_user
                firebase_user = get_firebase_user_by_email(user.email)
                if firebase_user:
                    # Actualizar solo en Firebase
                    update_result = update_firebase_user(firebase_user.uid, password=new_password)
                    if update_result:
                        # NO actualizar la contraseña en Django - solo usar Firebase
                        # user.set_password(new_password)  # COMENTADO - solo Firebase
                        # user.save()  # COMENTADO - solo Firebase
                        if is_own_password:
                            messages.success(request, 'Tu contraseña ha sido actualizada en Firebase exitosamente.')
                        else:
                            messages.success(request, f'Contraseña de {user.username} actualizada en Firebase exitosamente.')
                    else:
                        if is_own_password:
                            messages.error(request, 'Error al actualizar la contraseña en Firebase.')
                        else:
                            messages.error(request, f'Error al actualizar la contraseña de {user.username} en Firebase.')
                else:
                    if is_own_password:
                        messages.error(request, 'Usuario no encontrado en Firebase. No se pudo actualizar la contraseña.')
                    else:
                        messages.error(request, f'Usuario {user.username} no encontrado en Firebase. No se pudo actualizar la contraseña.')
            else:
                if is_own_password:
                    messages.error(request, 'El usuario no tiene email. No se puede actualizar la contraseña en Firebase.')
                else:
                    messages.error(request, f'El usuario {user.username} no tiene email. No se puede actualizar la contraseña en Firebase.')
            
            if is_own_password:
                return redirect('product_list')
            else:
                return redirect('user_list')
    else:
        form = UserPasswordChangeForm(user=user, is_own_password=is_own_password)
    
    return render(request, 'gestion/user_password_form.html', {
        'form': form,
        'user': user,
        'is_own_password': is_own_password
    })


@user_passes_test(is_admin, login_url='login')
def user_delete(request, pk):
    """
    Vista para eliminar un usuario.
    """
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        username = user.username
        email = user.email
        
        # Eliminar de Firebase si existe
        if email:
            firebase_user = get_firebase_user_by_email(email)
            if firebase_user:
                if delete_firebase_user(firebase_user.uid):
                    messages.success(request, f'Usuario {username} eliminado de Django y Firebase exitosamente.')
                else:
                    messages.warning(request, f'Usuario {username} eliminado de Django, pero no se pudo eliminar de Firebase.')
            else:
                messages.success(request, f'Usuario {username} eliminado exitosamente.')
        else:
            messages.success(request, f'Usuario {username} eliminado exitosamente.')
        
        user.delete()
        return redirect('user_list')
    
    return render(request, 'gestion/user_confirm_delete.html', {
        'user': user
    })


# ----------------------------------------------------
# VISTA AJAX PARA LIVE SEARCH DE USUARIOS
# ----------------------------------------------------
@login_required
def search_users_live(request):
    """
    Busca usuarios por 'q' y devuelve el HTML de la tabla.
    """
    if not is_admin(request.user):
        return JsonResponse({'html': '<tr><td colspan="7" class="text-center text-danger">No tienes permisos.</td></tr>'})
    
    search_query = request.GET.get('q', '')
    
    queryset = User.objects.all().select_related('profile')
    if search_query:
        queryset = queryset.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
    
    # Ordenamiento
    sort_by = request.GET.get('sort', 'username')
    sort_options = {
        'username': 'username',
        '-username': '-username',
        'email': 'email',
        '-email': '-email',
        'first_name': 'first_name',
        '-first_name': '-first_name',
        'last_name': 'last_name',
        '-last_name': '-last_name',
    }
    
    if sort_by in sort_options:
        queryset = queryset.order_by(sort_options[sort_by])
    else:
        queryset = queryset.order_by('username')
    
    context = {
        'users': queryset,
    }

    html_results = render_to_string(
        'gestion/includes/user_table_body.html', 
        context, 
        request=request
    )
    return JsonResponse({'html': html_results})

